# -*- coding: utf-8 -*-
'''
@File     : operExcel.py
@Copyright: Qiuwenjing
@Date     : 2021/10/14
@Desc     :
'''
import openpyxl
from openpyxl import load_workbook

def create_excel(filepath):
    f = openpyxl.Workbook()
    sheet1 = f.active
    f.save(filepath)

def open_Excel(filepath):
    wb = load_workbook(filepath)
    return wb

def get_sheet(wb,sheetName="Sheet1"):
    sheetNames = wb.sheetnames
    if sheetName not in sheetNames:
        raise IndexError('输入的sheet未找到，请检查%s' % sheetNames)
    ws = wb[sheetName]
    return ws

def get_rows(ws):
    rows = ws.rows
    return rows

def switch_to_CN(worksheet, rows, CN_data={}, insert_col=1):
    """插入首列，字段转换为中文
       @:CN_data 中英文对照字典
    """
    ws = worksheet.insert_cols(insert_col)
    '''for r in range(len(rows)):
        r[insert_col-1].value'''

def save_workbook(wb,filepath):
    wb.save(filepath)

